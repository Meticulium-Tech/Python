
from openpyxl import load_workbook
#this line permits the script to use the "load_workbook" class
#comming from the "openpyxl" library which is a library that permits the usage and
# manipulation of xlsx and csv files

wb = load_workbook('employeedata.xlsx')
#in this line, we are simply loading our existing excel file. Since the file is in the same directory 
# with the script, we simply input the name if not you copy the path of its location

sheet = wb['Sheet1']
#in this line, we simply declare that the workbook, containing the sheet named Sheet1, should
#be saved as sheet

for i in range(2, sheet.max_row+1):
    cell = sheet.cell(i, 2)
#we then declare a for loop that will iterate through the cells in rows found on the sheet
   
    if 'helpinghands.cm' in cell.value:
        update = (cell.value).replace('helpinghands.cm','handsinhands.org')
        sheet.cell(i,2).value = update
#the if condition comes in cause while going through the sheet, if the script finds the 'helpinghands.cm' in
#any of the cells, it will update it and replace it by 'handsinhands.org'
                
wb.save('updated_employeedata.csv')
wb.save('updated_employeedata.xlsx')
#Finally after updating and replacing it the script creates a new excel sheet and saves it using the
#workbook class and save attribute
